from datetime import date
import calendar
import smtplib
import openpyxl
from openpyxl import Workbook
from openpyxl.cell import Cell
import unicodedata

'''
	This file contains:
	- findlistName: find list PIC of assets
	- findPIC: find column in test sheet has name is "PIC"
	- findBarCode: find column in test sheet has name is "BarCode"
	- findAssetType: find column in test sheet has name is "AssetType"
	- findDescription: find column in test sheet has name is "Description"
	- findName_BarCode: find list Assets with input is PIC
'''

def findlistName(sheet):
	list = []
	columnPIC = findPIC(sheet)
	list.append(sheet.cell(row=3,column=columnPIC).value)
	for i in range (4,sheet.max_row):
		if sheet.cell(row=i,column=columnPIC).value != None:
			check=0
			for item in list:
				if(item.lower() == sheet.cell(row=i,column=columnPIC).value.lower()):
					check=1
			if check==0:
				list.append(sheet.cell(row=i,column=columnPIC).value)
	return list

'''
	This method will find column in sheet has name "PIC"
'''
def findPIC(sheet):
	for i in range (1,sheet.max_row):
		for j in range (1,sheet.max_column):
			if sheet.cell(row=i, column=j).value=="PIC":
				return j
'''
	This method will find column in sheet has name "Barcode"
'''
def findBarCode(sheet):
	for i in range (1,sheet.max_row):
		for j in range (1,sheet.max_column):
			if sheet.cell(row=i, column=j).value=="Barcode":
				return j
'''
	This method will find column in sheet has name "AssetType"
'''
def findAssetType(sheet):
	for i in range (1,sheet.max_row):
		for j in range (1,sheet.max_column):
			if sheet.cell(row=i, column=j).value=="AssetType":
				return j
'''
	This method will find column in sheet has name "Description"
'''
def findDescription(sheet):
	for i in range (1,sheet.max_row):
		for j in range (1,sheet.max_column):
			if sheet.cell(row=i, column=j).value=="Description":
				return j
'''
	This method will return list Barcode + AssetType + Description 
		1. Find Barcode, AssetType, Description with name
		2. Create a dict and save Barcode, AssetType, Description to the dict
		3. Create a list and add dict to the list
'''
def findName_BarCode(sheet,name):
	list=[]
	columnPIC = findPIC(sheet)
	columnAssetType = findAssetType(sheet)
	columnDescription = findDescription(sheet)
	columnBarCode=findBarCode(sheet)
	for i in range (3,sheet.max_row+1):
		if(sheet.cell(row=i,column=columnPIC).value != None):
			if(sheet.cell(row=i,column=columnPIC).value.lower()==name.lower()):
				dict = {}
				dict['Barcode'] = sheet.cell(row=i,column=columnBarCode).value
				dict['AssetType'] = sheet.cell(row=i,column=columnAssetType).value
				dict['Description'] = sheet.cell(row=i,column=columnDescription).value
				list.append(dict)
	return list
